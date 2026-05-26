"""
Normalize Large Load Economic Development Report Excel workbooks and Q1 2026 CSVs
into two combined datasets:
  - pipeline_snapshot.csv : total MW by (report_quarter, project_stage, planning_year)
  - pipeline_changes.csv  : per-quarter change metrics (additions, removals, etc.)
"""
import re
import sys
from pathlib import Path
import openpyxl
import pandas as pd

SCRIPT_DIR = Path(__file__).parent
OUTPUTS_DIR = SCRIPT_DIR.parent / "outputs"
WORKBOOKS_DIR = OUTPUTS_DIR / "workbooks"
CSV_2026Q1_DIR = OUTPUTS_DIR / "2026Q1"
COMBINED_DIR = OUTPUTS_DIR / "combined"

STAGE_ALIASES = {
    "contract for electric service": "Contract for Electric Service",
    "request for service": "Request for Service",
    "request for electric service": "Request for Service",
    "technical review": "Technical Review",
    "aggregate total": "Aggregate Total",
}

QUARTER_ORDER = ["2024Q1","2024Q2","2024Q3","2024Q4",
                 "2025Q1","2025Q2","2025Q3","2025Q4","2026Q1"]


def normalize_stage(raw: str) -> str:
    if not raw:
        return ""
    cleaned = re.sub(r'[\s\d*]+$', '', str(raw)).strip()
    return STAGE_ALIASES.get(cleaned.lower(), cleaned)


def load_sheet_rows(ws) -> list[tuple]:
    """Load all rows from a worksheet in one pass (safe for read_only mode)."""
    return list(ws.iter_rows(values_only=True))


def find_header_idx(rows: list[tuple]) -> int | None:
    """Return 0-indexed position of the row containing 'Project Name' or 'Project Stage'."""
    for i, row in enumerate(rows):
        for cell in row:
            if cell and re.search(r'project (name|stage)', str(cell), re.IGNORECASE):
                return i
    return None


def get_year_cols(headers: tuple) -> dict[int, int]:
    """Return {col_index: year} for columns whose header looks like a 4-digit year."""
    result = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        m = re.fullmatch(r'(\d{4})', str(h).strip())
        if m:
            yr = int(m.group(1))
            if 2020 <= yr <= 2040:
                result[i] = yr
    return result


def col_index(headers: list, *names: str) -> int | None:
    """Find first matching column (case-insensitive, partial match allowed)."""
    for name in names:
        for i, h in enumerate(headers):
            if h and name.lower() in str(h).lower():
                return i
    return None


# ---------------------------------------------------------------------------
# Excel parsers
# ---------------------------------------------------------------------------

def parse_main_sheet(ws, quarter: str) -> list[dict]:
    """Extract project-level rows from the Main sheet."""
    rows = load_sheet_rows(ws)
    hrow = find_header_idx(rows)
    if hrow is None:
        return []
    headers = rows[hrow]
    year_cols = get_year_cols(headers)
    stage_col = col_index(headers, "project stage")
    if stage_col is None:
        print(f"    WARNING: 'Project Stage' column not found in {quarter} Main sheet")
        return []

    records = []
    for row in rows[hrow + 1:]:
        if not any(c is not None for c in row):
            continue
        raw_stage = row[stage_col]
        if raw_stage is None:
            continue
        stage = normalize_stage(raw_stage)
        if not stage:
            continue
        for ci, year in year_cols.items():
            val = row[ci]
            try:
                mw = float(val) if val is not None else 0.0
            except (TypeError, ValueError):
                mw = 0.0
            records.append({
                "report_quarter": quarter,
                "project_stage": stage,
                "planning_year": year,
                "load_mw": mw,
            })
    return records


def parse_removed_projects(ws) -> tuple[int, float]:
    """Return (count, total_mw) of removed projects."""
    rows = load_sheet_rows(ws)
    hrow = find_header_idx(rows)
    if hrow is None:
        return 0, 0.0
    headers = rows[hrow]
    mw_col = col_index(headers, "announced load")
    if mw_col is None:
        return 0, 0.0
    count, total = 0, 0.0
    for row in rows[hrow + 1:]:
        if not any(c is not None for c in row):
            continue
        if row[0] is None:
            continue
        try:
            total += float(row[mw_col]) if row[mw_col] is not None else 0.0
        except (TypeError, ValueError):
            pass
        count += 1
    return count, total


def parse_projects_added(ws) -> tuple[int, float]:
    """Return (count, total_mw) of projects added."""
    rows = load_sheet_rows(ws)
    hrow = find_header_idx(rows)
    if hrow is None:
        return 0, 0.0
    headers = rows[hrow]
    mw_col = col_index(headers, "announced load")
    if mw_col is None:
        return 0, 0.0
    count, total = 0, 0.0
    for row in rows[hrow + 1:]:
        if not any(c is not None for c in row):
            continue
        if row[0] is None:
            continue
        try:
            total += float(row[mw_col]) if row[mw_col] is not None else 0.0
        except (TypeError, ValueError):
            pass
        count += 1
    return count, total


def parse_load_change(ws) -> float:
    """Return net MW change from Load Change sheet (sum of Change column)."""
    rows = load_sheet_rows(ws)
    hrow = find_header_idx(rows)
    if hrow is None:
        return 0.0
    headers = rows[hrow]
    change_col = col_index(headers, "change")
    if change_col is None:
        return 0.0
    total = 0.0
    for row in rows[hrow + 1:]:
        if not any(c is not None for c in row):
            continue
        if row[0] is None:
            continue
        try:
            total += float(row[change_col]) if row[change_col] is not None else 0.0
        except (TypeError, ValueError):
            pass
    return total


def parse_service_date_change(ws) -> float:
    """Return average delay in months from Initial Service Change sheet."""
    rows = load_sheet_rows(ws)
    hrow = find_header_idx(rows)
    if hrow is None:
        return 0.0
    headers = rows[hrow]
    change_col = col_index(headers, "change (months)", "change")
    if change_col is None:
        return 0.0
    vals = []
    for row in rows[hrow + 1:]:
        if not any(c is not None for c in row):
            continue
        if row[0] is None:
            continue
        try:
            v = float(row[change_col]) if row[change_col] is not None else None
            if v is not None:
                vals.append(v)
        except (TypeError, ValueError):
            pass
    return sum(vals) / len(vals) if vals else 0.0


def parse_stage_change(ws) -> int:
    """Return count of projects that changed stage."""
    rows = load_sheet_rows(ws)
    hrow = find_header_idx(rows)
    if hrow is None:
        return 0
    count = 0
    for row in rows[hrow + 1:]:
        if row[0] is not None and any(c is not None for c in row):
            count += 1
    return count


def process_excel(xlsx_path: Path, quarter: str) -> tuple[list[dict], dict]:
    """Parse all relevant sheets from one Excel. Returns (snapshot_rows, change_metrics)."""
    print(f"  Parsing {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = {s.lower(): s for s in wb.sheetnames}

    snapshot_rows = []
    if "main" in sheet_names:
        snapshot_rows = parse_main_sheet(wb[sheet_names["main"]], quarter)
        print(f"    Main: {len(snapshot_rows)} project-year rows")
    else:
        print(f"    WARNING: no 'Main' sheet found")

    removed_count, removed_mw = 0, 0.0
    if "removed projects" in sheet_names:
        removed_count, removed_mw = parse_removed_projects(wb[sheet_names["removed projects"]])

    added_count, added_mw = 0, 0.0
    if "projects added" in sheet_names:
        added_count, added_mw = parse_projects_added(wb[sheet_names["projects added"]])

    load_change_net = 0.0
    if "load change" in sheet_names:
        load_change_net = parse_load_change(wb[sheet_names["load change"]])

    avg_delay = 0.0
    if "initial service change" in sheet_names:
        avg_delay = parse_service_date_change(wb[sheet_names["initial service change"]])

    stage_changes = 0
    if "stage change" in sheet_names:
        stage_changes = parse_stage_change(wb[sheet_names["stage change"]])

    change_metrics = {
        "report_quarter": quarter,
        "added_projects": added_count,
        "added_mw": added_mw,
        "removed_projects": removed_count,
        "removed_mw": removed_mw,
        "load_change_net_mw": load_change_net,
        "stage_changes_count": stage_changes,
        "avg_delay_months": round(avg_delay, 2),
    }
    return snapshot_rows, change_metrics


# ---------------------------------------------------------------------------
# Q1 2026 CSV ingestion
# ---------------------------------------------------------------------------

def process_2026q1_csvs() -> tuple[list[dict], dict]:
    quarter = "2026Q1"
    print(f"  Reading Q1 2026 CSVs from {CSV_2026Q1_DIR}")

    df_main = pd.read_csv(CSV_2026Q1_DIR / "pipeline_main.csv")
    # Year columns may be "2026" or "Load_2026" — detect both patterns
    year_cols = {}  # col_name -> year int
    for c in df_main.columns:
        m = re.search(r'(\d{4})$', c)
        if m:
            yr = int(m.group(1))
            if 2020 <= yr <= 2040:
                year_cols[c] = yr
    stage_col = next((c for c in df_main.columns if "stage" in c.lower()), None)
    if stage_col is None:
        print("    WARNING: no stage column in pipeline_main.csv")
        return [], {}

    snapshot_rows = []
    for _, row in df_main.iterrows():
        stage = normalize_stage(row[stage_col])
        if not stage:
            continue
        for yr_col, yr in year_cols.items():
            try:
                mw = float(row[yr_col]) if pd.notna(row[yr_col]) else 0.0
            except (TypeError, ValueError):
                mw = 0.0
            snapshot_rows.append({
                "report_quarter": quarter,
                "project_stage": stage,
                "planning_year": yr,
                "load_mw": mw,
            })
    print(f"    Main: {len(snapshot_rows)} project-year rows")

    def _count_mw(filename, mw_col_hint):
        path = CSV_2026Q1_DIR / filename
        if not path.exists():
            return 0, 0.0
        df = pd.read_csv(path)
        mw_col = next((c for c in df.columns if mw_col_hint.lower() in c.lower()), None)
        count = len(df.dropna(subset=[df.columns[0]]))
        total = df[mw_col].apply(pd.to_numeric, errors='coerce').sum() if mw_col else 0.0
        return count, float(total)

    added_count, added_mw = _count_mw("new_projects.csv", "load")
    removed_count, removed_mw = _count_mw("pipeline_exits.csv", "load")

    load_change_net = 0.0
    lc_path = CSV_2026Q1_DIR / "load_changes.csv"
    if lc_path.exists():
        df_lc = pd.read_csv(lc_path)
        change_col = next((c for c in df_lc.columns if "change" in c.lower()), None)
        if change_col:
            load_change_net = float(df_lc[change_col].apply(pd.to_numeric, errors='coerce').sum())

    avg_delay = 0.0
    sd_path = CSV_2026Q1_DIR / "service_date_changes.csv"
    if sd_path.exists():
        df_sd = pd.read_csv(sd_path)
        delay_col = next((c for c in df_sd.columns if "month" in c.lower() or "change" in c.lower()), None)
        if delay_col:
            vals = df_sd[delay_col].apply(pd.to_numeric, errors='coerce').dropna()
            avg_delay = float(vals.mean()) if len(vals) else 0.0

    stage_changes = 0
    sc_path = CSV_2026Q1_DIR / "stage_changes.csv"
    if sc_path.exists():
        df_sc = pd.read_csv(sc_path)
        stage_changes = len(df_sc.dropna(subset=[df_sc.columns[0]]))

    change_metrics = {
        "report_quarter": quarter,
        "added_projects": added_count,
        "added_mw": added_mw,
        "removed_projects": removed_count,
        "removed_mw": removed_mw,
        "load_change_net_mw": load_change_net,
        "stage_changes_count": stage_changes,
        "avg_delay_months": round(avg_delay, 2),
    }
    return snapshot_rows, change_metrics


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    COMBINED_DIR.mkdir(parents=True, exist_ok=True)

    all_snapshot_rows: list[dict] = []
    all_change_metrics: list[dict] = []

    # Process Excel workbooks
    quarter_dirs = sorted(
        [d for d in WORKBOOKS_DIR.iterdir() if d.is_dir() and d.name != "raw"],
        key=lambda d: QUARTER_ORDER.index(d.name) if d.name in QUARTER_ORDER else 99
    )

    for qdir in quarter_dirs:
        quarter = qdir.name
        xlsxs = list(qdir.glob("*.xlsx"))
        if not xlsxs:
            print(f"\n[{quarter}] No Excel files found, skipping")
            continue
        print(f"\n[{quarter}]")
        # There should be exactly one Excel per quarter dir
        snap_rows, change_metrics = process_excel(xlsxs[0], quarter)
        all_snapshot_rows.extend(snap_rows)
        all_change_metrics.append(change_metrics)

    # Process Q1 2026 CSVs
    print(f"\n[2026Q1]")
    snap_rows, change_metrics = process_2026q1_csvs()
    all_snapshot_rows.extend(snap_rows)
    all_change_metrics.append(change_metrics)

    # Build pipeline_snapshot.csv — aggregate by (quarter, stage, year)
    df_snap = pd.DataFrame(all_snapshot_rows)
    df_snap = (
        df_snap
        .groupby(["report_quarter", "project_stage", "planning_year"], as_index=False)
        ["load_mw"].sum()
    )
    df_snap["planning_year"] = df_snap["planning_year"].astype(int)
    df_snap = df_snap.sort_values(["report_quarter", "project_stage", "planning_year"])

    snap_path = COMBINED_DIR / "pipeline_snapshot.csv"
    df_snap.to_csv(snap_path, index=False)
    print(f"\nWrote {snap_path} ({len(df_snap)} rows)")

    # Build pipeline_changes.csv
    df_changes = pd.DataFrame(all_change_metrics)
    df_changes = df_changes.set_index("report_quarter").reindex(QUARTER_ORDER).reset_index()
    changes_path = COMBINED_DIR / "pipeline_changes.csv"
    df_changes.to_csv(changes_path, index=False)
    print(f"Wrote {changes_path} ({len(df_changes)} rows)")

    # Summary table
    print("\n--- Per-quarter summary ---")
    print(f"{'Quarter':<10} {'Projects':>8} {'Stages':<50} {'Total MW (2030)':>14}")
    print("-" * 85)
    for q in QUARTER_ORDER:
        sub = df_snap[df_snap.report_quarter == q]
        if sub.empty:
            print(f"{q:<10}  (no data)")
            continue
        total_project_rows = sub.load_mw.count()
        stages = ", ".join(sorted(sub.project_stage.unique()))
        mw_2030 = sub[sub.planning_year == 2030].load_mw.sum()
        print(f"{q:<10} {total_project_rows:>8} {stages:<50} {mw_2030:>14,.0f}")


if __name__ == "__main__":
    main()
