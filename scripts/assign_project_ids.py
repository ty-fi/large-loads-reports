"""
assign_project_ids.py

Tracks individual large load projects across quarterly GA PSC Docket 55378 reports.
Assigns stable project IDs using fingerprint matching (segment, territory, announced_load,
initial_service_date) augmented by load ramp cosine similarity as a tiebreaker.

Quarter-to-quarter update logic:
  1. Apply removals from Removed Projects sheet (match by prev announced_load)
  2. Apply load changes from Load Change sheet (update announced_load in master)
  3. Apply date changes from Initial Service Change sheet (update ISD in master)
  4. Match current Main sheet projects against updated fingerprints
     - Pass 1: exact match (with segment)
     - Pass 2: exact match ignoring segment (Q1 2024 → Q1 2025 format transition)
     - Pass 3: ramp cosine similarity >= 0.70 among still-unmatched candidates
     - Remainder: new projects, assigned next sequential ID

Output: outputs/combined/pipeline_projects.csv
"""

import re
from copy import deepcopy
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

SCRIPT_DIR = Path(__file__).parent
OUTPUTS_DIR = SCRIPT_DIR.parent / "outputs"
WORKBOOKS_DIR = OUTPUTS_DIR / "workbooks"
CSV_2026Q1_DIR = OUTPUTS_DIR / "2026Q1"
COMBINED_DIR = OUTPUTS_DIR / "combined"

QUARTER_ORDER = ["2024Q1", "2024Q2", "2024Q3", "2024Q4",
                 "2025Q1", "2025Q2", "2025Q3", "2025Q4", "2026Q1"]

LOAD_YEARS = list(range(2023, 2038))

RAMP_SIM_THRESHOLD = 0.70

STAGE_ALIASES = {
    "contract for electric service": "Contract for Electric Service",
    "request for service": "Request for Service",
    "request for electric service": "Request for Service",
    "technical review": "Technical Review",
}

SEGMENT_ALIASES = {
    "data centers": "data center",
    "data center/crypto": "data center",
    "clean energy technology": "clean energy tech",
}


# ── Normalization ─────────────────────────────────────────────────────────────

def norm_stage(raw):
    if not raw:
        return ""
    cleaned = re.sub(r'[\s\d*]+$', '', str(raw)).strip()
    return STAGE_ALIASES.get(cleaned.lower(), cleaned)

def norm_segment(s):
    if not s:
        return ""
    return SEGMENT_ALIASES.get(s.strip().lower(), s.strip().lower())

def norm_load(v):
    try:
        return int(round(float(v))) if v is not None else None
    except (TypeError, ValueError):
        return None

def norm_date(s):
    if not s:
        return ""
    m = re.search(r'Q([1-4])\s*(\d{4})', str(s), re.IGNORECASE)
    return f"Q{m.group(1)} {m.group(2)}" if m else str(s).strip()

def norm_flag(v):
    return "Y" if str(v).strip().upper() == "Y" else ""


# ── Low-level sheet helpers ───────────────────────────────────────────────────

def load_ws_rows(ws):
    return list(ws.iter_rows(values_only=True))

def find_header_row(rows):
    for i, row in enumerate(rows):
        if any(c and re.search(r'project.?(name|stage)', str(c), re.IGNORECASE) for c in row):
            return i
    return None

def find_col(headers, *patterns):
    for pat in patterns:
        for i, h in enumerate(headers):
            if h and re.search(pat, str(h), re.IGNORECASE):
                return i
    return None


# ── Excel Main sheet parser ───────────────────────────────────────────────────

def parse_excel_main(ws, quarter):
    rows = load_ws_rows(ws)
    hi = find_header_row(rows)
    if hi is None:
        return []
    headers = rows[hi]

    # Year columns (bare 4-digit integers/strings)
    year_cols = {}
    for i, h in enumerate(headers):
        m = re.fullmatch(r'(\d{4})', str(h).strip()) if h is not None else None
        if m:
            yr = int(m.group(1))
            if 2020 <= yr <= 2040:
                year_cols[i] = yr

    # Change flag columns appear after the last year column in a fixed order:
    # New Project?, Announced Load (chg), Load Ramp (chg), Project Stage (chg), Initial Service Date (chg)
    last_yr_idx = max(year_cols.keys()) if year_cols else -1
    flag_positions = [i for i, h in enumerate(headers) if i > last_yr_idx and h is not None]
    flag_keys = ['new_project', 'change_announced_load', 'change_load_ramp',
                 'change_project_stage', 'change_initial_service_date']
    flag_col = {k: flag_positions[i] if i < len(flag_positions) else None
                for i, k in enumerate(flag_keys)}

    seg_col   = find_col(headers, r'^segment$')
    class_col = find_col(headers, r'^class$', r'facility.?type')
    terr_col  = find_col(headers, r'^territory$')
    stage_col = find_col(headers, r'project.?stage')
    load_col  = find_col(headers, r'announced.?load')
    date_col  = find_col(headers, r'in.?service.?date', r'initial.?service')

    projects = []
    for row in rows[hi + 1:]:
        if not any(c is not None for c in row):
            continue
        stage_val = row[stage_col] if stage_col is not None else None
        if stage_val is None:
            continue
        stage = norm_stage(stage_val)
        if not stage:
            continue

        seg   = norm_segment(str(row[seg_col]).strip() if seg_col is not None and row[seg_col] else "")
        cls   = str(row[class_col]).strip() if class_col is not None and row[class_col] else ""
        terr  = str(row[terr_col]).strip() if terr_col is not None and row[terr_col] else ""
        load  = norm_load(row[load_col] if load_col is not None else None)
        date  = norm_date(row[date_col] if date_col is not None else None)

        ramp = {}
        for ci, yr in year_cols.items():
            v = row[ci]
            try:
                ramp[yr] = float(v) if v is not None else 0.0
            except (TypeError, ValueError):
                ramp[yr] = 0.0

        def get_flag(k):
            idx = flag_col.get(k)
            if idx is None:
                return ""
            return norm_flag(row[idx]) if row[idx] is not None else ""

        projects.append({
            'segment': seg, 'class_type': cls, 'territory': terr,
            'project_stage': stage, 'announced_load': load,
            'initial_service_date': date, 'ramp': ramp,
            'new_project':                   get_flag('new_project'),
            'change_announced_load':         get_flag('change_announced_load'),
            'change_load_ramp':              get_flag('change_load_ramp'),
            'change_project_stage':          get_flag('change_project_stage'),
            'change_initial_service_date':   get_flag('change_initial_service_date'),
        })
    return projects


# ── Excel change sheet parsers ────────────────────────────────────────────────

def parse_excel_load_changes(ws):
    """Return list of (prev_load, new_load)."""
    rows = load_ws_rows(ws)
    hi = find_header_row(rows)
    if hi is None:
        return []
    result = []
    for row in rows[hi + 1:]:
        if row[0] is None:
            continue
        new_load  = norm_load(row[1])
        prev_load = norm_load(row[2])
        if prev_load is not None and new_load is not None and prev_load != new_load:
            result.append((prev_load, new_load))
    return result

def parse_excel_date_changes(ws):
    """Return list of (prev_date, new_date)."""
    rows = load_ws_rows(ws)
    hi = find_header_row(rows)
    if hi is None:
        return []
    result = []
    for row in rows[hi + 1:]:
        if row[0] is None:
            continue
        new_date  = norm_date(row[1])
        prev_date = norm_date(row[2])
        if prev_date and new_date and prev_date != new_date:
            result.append((prev_date, new_date))
    return result

def parse_excel_ramp_changes(ws):
    """Return list of (prev_ramp_dict, new_ramp_dict).

    Sheet layout (row 0 = section labels, row 1 = year headers):
      col 0          : Project Name
      cols 1-15      : new ramp (current quarter), years 2023-2037
      cols 16-30     : previous ramp, years 2023-2037
    """
    rows = load_ws_rows(ws)
    # Header is row 1 (year integers); row 0 has section labels
    hi = None
    for i, row in enumerate(rows):
        if any(isinstance(c, int) and 2020 <= c <= 2040 for c in row):
            hi = i
            break
    if hi is None:
        return []

    # Identify which columns are year columns in the first block (new ramp)
    # and second block (prev ramp)
    year_new, year_prev = {}, {}
    seen_years = set()
    for ci, h in enumerate(rows[hi]):
        if isinstance(h, int) and 2020 <= h <= 2040:
            yr = h
            if yr not in seen_years:
                year_new[ci] = yr
                seen_years.add(yr)
            elif yr not in {v for v in year_prev.values()}:
                year_prev[ci] = yr

    if not year_new or not year_prev:
        return []

    result = []
    for row in rows[hi + 1:]:
        if not row or row[0] is None:
            continue
        new_ramp  = {yr: float(row[ci]) if row[ci] is not None else 0.0
                     for ci, yr in year_new.items()}
        prev_ramp = {yr: float(row[ci]) if row[ci] is not None else 0.0
                     for ci, yr in year_prev.items()}
        if any(v != 0 for v in new_ramp.values()) or any(v != 0 for v in prev_ramp.values()):
            result.append((prev_ramp, new_ramp))
    return result


def parse_excel_removed(ws):
    """Return list of (prev_announced_load, reason)."""
    rows = load_ws_rows(ws)
    hi = find_header_row(rows)
    if hi is None:
        return []
    result = []
    for row in rows[hi + 1:]:
        if row[0] is None:
            continue
        load   = norm_load(row[1])
        reason = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if load is not None:
            result.append((load, reason))
    return result


# ── Q1 2026 CSV parsers ───────────────────────────────────────────────────────

def parse_csv_main(quarter):
    df = pd.read_csv(CSV_2026Q1_DIR / "pipeline_main.csv")
    year_cols = {c: int(re.search(r'(\d{4})$', c).group(1))
                 for c in df.columns if re.search(r'(\d{4})$', c) and
                 2020 <= int(re.search(r'(\d{4})$', c).group(1)) <= 2040}

    stage_col = next((c for c in df.columns if 'stage' in c.lower()), None)
    load_col  = next((c for c in df.columns if 'announced_load' in c.lower()), None)
    date_col  = next((c for c in df.columns if 'service_date' in c.lower()), None)
    seg_col   = next((c for c in df.columns if c.lower() == 'segment'), None)
    class_col = next((c for c in df.columns if 'facility_type' in c.lower() or c.lower() == 'class'), None)
    terr_col  = next((c for c in df.columns if 'territory' in c.lower()), None)

    flag_map = {
        'new_project':                 next((c for c in df.columns if 'new_project' in c.lower()), None),
        'change_announced_load':       next((c for c in df.columns if 'change_announced' in c.lower()), None),
        'change_load_ramp':            next((c for c in df.columns if 'change_load_ramp' in c.lower()), None),
        'change_project_stage':        next((c for c in df.columns if 'change_project_stage' in c.lower()), None),
        'change_initial_service_date': next((c for c in df.columns if 'change_initial' in c.lower()), None),
    }

    projects = []
    for _, row in df.iterrows():
        stage = norm_stage(row[stage_col] if stage_col else None)
        if not stage:
            continue

        ramp = {}
        for col, yr in year_cols.items():
            try:
                ramp[yr] = float(row[col]) if pd.notna(row[col]) else 0.0
            except (TypeError, ValueError):
                ramp[yr] = 0.0

        def get_flag(k):
            col = flag_map.get(k)
            if col is None:
                return ""
            v = row[col]
            return norm_flag(v) if pd.notna(v) else ""

        projects.append({
            'segment':    norm_segment(str(row[seg_col]).strip() if seg_col and pd.notna(row[seg_col]) else ""),
            'class_type': str(row[class_col]).strip() if class_col and pd.notna(row[class_col]) else "",
            'territory':  str(row[terr_col]).strip() if terr_col and pd.notna(row[terr_col]) else "",
            'project_stage': stage,
            'announced_load': norm_load(row[load_col] if load_col else None),
            'initial_service_date': norm_date(row[date_col] if date_col else None),
            'ramp': ramp,
            'new_project':                   get_flag('new_project'),
            'change_announced_load':         get_flag('change_announced_load'),
            'change_load_ramp':              get_flag('change_load_ramp'),
            'change_project_stage':          get_flag('change_project_stage'),
            'change_initial_service_date':   get_flag('change_initial_service_date'),
        })
    return projects

def parse_csv_load_changes():
    df = pd.read_csv(CSV_2026Q1_DIR / "load_changes.csv")
    result = []
    prev_col = next((c for c in df.columns if 'q4_2025' in c.lower() and 'load' in c.lower()), None)
    new_col  = next((c for c in df.columns if 'q1_2026' in c.lower() and 'load' in c.lower()), None)
    if prev_col and new_col:
        for _, row in df.iterrows():
            p, n = norm_load(row[prev_col]), norm_load(row[new_col])
            if p is not None and n is not None and p != n:
                result.append((p, n))
    return result

def parse_csv_date_changes():
    df = pd.read_csv(CSV_2026Q1_DIR / "service_date_changes.csv")
    result = []
    prev_col = next((c for c in df.columns if 'q4_2025' in c.lower() and 'date' in c.lower()), None)
    new_col  = next((c for c in df.columns if 'q1_2026' in c.lower() and 'date' in c.lower()), None)
    if prev_col and new_col:
        for _, row in df.iterrows():
            p, n = norm_date(row[prev_col]), norm_date(row[new_col])
            if p and n and p != n:
                result.append((p, n))
    return result

def parse_csv_removed():
    df = pd.read_csv(CSV_2026Q1_DIR / "pipeline_exits.csv")
    load_col   = next((c for c in df.columns if 'announced_load' in c.lower()), None)
    reason_col = next((c for c in df.columns if 'reason' in c.lower()), None)
    result = []
    if load_col:
        for _, row in df.iterrows():
            load   = norm_load(row[load_col])
            reason = str(row[reason_col]).strip() if reason_col and pd.notna(row[reason_col]) else ""
            if load is not None:
                result.append((load, reason))
    return result


# ── Fingerprint & similarity ──────────────────────────────────────────────────

def make_fp(seg, terr, load, date, use_seg=True):
    s = seg if (use_seg and seg) else ""
    return (s, terr.lower(), load, date.upper() if date else "")

def ramp_sim(ramp1, ramp2):
    """Cosine similarity * magnitude ratio.
    The magnitude penalty ensures that flat ramps of different scales
    (e.g. 110 MW vs 180 MW) don't score 1.0 just because they're proportional.
    """
    v1 = np.array([float(ramp1.get(yr, 0)) for yr in LOAD_YEARS])
    v2 = np.array([float(ramp2.get(yr, 0)) for yr in LOAD_YEARS])
    n1, n2 = np.linalg.norm(v1), np.linalg.norm(v2)
    if n1 == 0 or n2 == 0:
        return 0.0
    cosine    = float(np.dot(v1, v2) / (n1 * n2))
    mag_ratio = min(n1, n2) / max(n1, n2)
    return cosine * mag_ratio


# ── Core matching ─────────────────────────────────────────────────────────────

def match_quarter(quarter, current_projects, prev_master, load_changes, date_changes,
                  ramp_changes, removed_loads, next_id):
    """
    Returns (result_rows, new_master, next_id).
    result_rows: one dict per current project with proj_id and all fields.
    new_master: updated {proj_id: record} for use in the next quarter.
    """
    working = deepcopy(prev_master)

    # ── Apply removals ────────────────────────────────────────────────────────
    # Count how many times each load appears in the removal list so we don't
    # over-remove when multiple master projects share the same announced_load.
    from collections import Counter
    removal_counts = Counter(rem_load for rem_load, _ in removed_loads)

    removed_pids = set()
    for rem_load, quota in removal_counts.items():
        hits = [pid for pid, r in working.items()
                if r['announced_load'] == rem_load and pid not in removed_pids]
        # Remove at most `quota` projects; prefer unambiguous cases
        for pid in hits[:quota]:
            removed_pids.add(pid)
    for pid in removed_pids:
        working.pop(pid, None)

    # ── Apply load changes ────────────────────────────────────────────────────
    for (prev_load, new_load) in load_changes:
        hits = [pid for pid in working if working[pid]['announced_load'] == prev_load]
        if len(hits) == 1:
            working[hits[0]]['announced_load'] = new_load
        # If 0 or >1: leave as-is; ramp matching will handle it

    # ── Apply date changes ────────────────────────────────────────────────────
    for (prev_date, new_date) in date_changes:
        hits = [pid for pid in working if working[pid]['initial_service_date'] == prev_date]
        if len(hits) == 1:
            working[hits[0]]['initial_service_date'] = new_date

    # ── Apply ramp changes ────────────────────────────────────────────────────
    # Match on exact previous ramp; update master ramp so Pass 3 similarity
    # compares against the most recent known ramp, not a stale one.
    for (prev_ramp, new_ramp) in ramp_changes:
        hits = [pid for pid in working
                if all(abs(working[pid]['ramp'].get(yr, 0) - prev_ramp.get(yr, 0)) < 0.5
                       for yr in LOAD_YEARS)]
        if len(hits) == 1:
            working[hits[0]]['ramp'] = new_ramp

    # ── Build fingerprint indexes ─────────────────────────────────────────────
    fp_idx      = {}  # (with seg) → [pid]
    fp_noseg_idx = {}  # (no seg)  → [pid]
    for pid, rec in working.items():
        fp  = make_fp(rec['segment'], rec['territory'], rec['announced_load'],
                      rec['initial_service_date'], use_seg=True)
        fpn = make_fp('', rec['territory'], rec['announced_load'],
                      rec['initial_service_date'], use_seg=False)
        fp_idx.setdefault(fp, []).append(pid)
        fp_noseg_idx.setdefault(fpn, []).append(pid)

    assigned = {}   # proj idx → proj_id
    confidence = {} # proj idx → str
    used = set()

    # ── Pass 1: exact match with segment ─────────────────────────────────────
    for i, proj in enumerate(current_projects):
        fp = make_fp(proj['segment'], proj['territory'], proj['announced_load'],
                     proj['initial_service_date'])
        hits = [p for p in fp_idx.get(fp, []) if p not in used]
        if len(hits) == 1:
            assigned[i] = hits[0]
            confidence[i] = 'exact'
            used.add(hits[0])

    # ── Pass 2: exact match ignoring segment (format transition) ──────────────
    # When multiple candidates share the same noseg fingerprint, tiebreak by
    # ramp similarity (handles Q1 duplicate projects with identical attributes).
    for i, proj in enumerate(current_projects):
        if i in assigned:
            continue
        fp = make_fp('', proj['territory'], proj['announced_load'],
                     proj['initial_service_date'], use_seg=False)
        hits = [p for p in fp_noseg_idx.get(fp, []) if p not in used]
        if len(hits) == 1:
            assigned[i] = hits[0]
            confidence[i] = 'exact_noseg'
            used.add(hits[0])
        elif len(hits) > 1:
            sims = sorted(
                [(ramp_sim(proj['ramp'], working[p]['ramp']), p) for p in hits],
                reverse=True,
            )
            best_s, best_p = sims[0]
            runner_s = sims[1][0] if len(sims) > 1 else 0.0
            # Require a clear ramp winner above threshold and a meaningful margin
            if best_s >= RAMP_SIM_THRESHOLD and best_s - runner_s >= 0.05:
                assigned[i] = best_p
                confidence[i] = 'exact_noseg'
                used.add(best_p)

    # ── Pass 3: ramp cosine similarity ────────────────────────────────────────
    unmatched = [i for i in range(len(current_projects)) if i not in assigned]
    available = [pid for pid in working if pid not in used]

    for i in unmatched:
        proj = current_projects[i]
        best_pid, best_sim = None, RAMP_SIM_THRESHOLD
        for pid in available:
            if pid in used:
                continue
            sim = ramp_sim(proj['ramp'], working[pid]['ramp'])
            if sim > best_sim:
                best_sim, best_pid = sim, pid
        if best_pid:
            assigned[i] = best_pid
            confidence[i] = f'ramp_{best_sim:.2f}'
            used.add(best_pid)

    # ── Pass 4: identity match (territory + load + date, no ramp threshold) ───
    # Catches revised-filing quarters where ramp data changed but other
    # identifying attributes are stable (e.g. 2025Q3 revised filing).
    identity_idx = {}
    for pid, rec in working.items():
        key = (rec['territory'].lower(), rec['announced_load'],
               rec['initial_service_date'].upper() if rec['initial_service_date'] else '')
        identity_idx.setdefault(key, []).append(pid)

    for i in [i for i in range(len(current_projects)) if i not in assigned]:
        proj = current_projects[i]
        key = (proj['territory'].lower(), proj['announced_load'],
               proj['initial_service_date'].upper() if proj['initial_service_date'] else '')
        hits = [p for p in identity_idx.get(key, []) if p not in used]
        if not hits:
            continue
        if len(hits) == 1:
            assigned[i] = hits[0]
            confidence[i] = 'identity'
            used.add(hits[0])
        else:
            # Tiebreak by best ramp sim; require a clear margin over runner-up
            sims = sorted(
                [(ramp_sim(proj['ramp'], working[p]['ramp']), p) for p in hits],
                reverse=True,
            )
            best_s, best_p = sims[0]
            runner_s = sims[1][0] if len(sims) > 1 else 0.0
            if best_s - runner_s >= 0.05:
                assigned[i] = best_p
                confidence[i] = f'identity_{best_s:.2f}'
                used.add(best_p)

    # ── Build output rows & new master ────────────────────────────────────────
    new_master = deepcopy(working)
    result_rows = []
    warnings = []

    for i, proj in enumerate(current_projects):
        if i in assigned:
            pid = assigned[i]
            conf = confidence[i]
        else:
            pid = f"P{next_id:04d}"
            next_id += 1
            conf = 'new'

        # Update master record with current quarter's state
        new_master[pid] = {
            'segment':              proj['segment'],
            'class_type':           proj['class_type'],
            'territory':            proj['territory'],
            'announced_load':       proj['announced_load'],
            'initial_service_date': proj['initial_service_date'],
            'ramp':                 proj['ramp'],
            'quarter_first_seen':   (prev_master[pid]['quarter_first_seen']
                                     if pid in prev_master else quarter),
        }

        row = {
            'proj_id':                       pid,
            'report_quarter':                quarter,
            'quarter_first_seen':            new_master[pid]['quarter_first_seen'],
            'match_confidence':              conf,
            'segment':                       proj['segment'],
            'class_type':                    proj['class_type'],
            'territory':                     proj['territory'],
            'project_stage':                 proj['project_stage'],
            'announced_load_mw':             proj['announced_load'],
            'initial_service_date':          proj['initial_service_date'],
            'new_project':                   proj['new_project'],
            'change_announced_load':         proj['change_announced_load'],
            'change_load_ramp':              proj['change_load_ramp'],
            'change_project_stage':          proj['change_project_stage'],
            'change_initial_service_date':   proj['change_initial_service_date'],
        }
        for yr in LOAD_YEARS:
            row[f'load_{yr}'] = proj['ramp'].get(yr) if yr in proj['ramp'] else None
        result_rows.append(row)

    # Warn about projects in working that were never matched
    unmatched_prev = [pid for pid in working if pid not in used]
    if unmatched_prev:
        warnings.append(f"  WARNING: {len(unmatched_prev)} prev-quarter project(s) not matched "
                        f"in {quarter} (may be removed or data gap): {unmatched_prev[:5]}")

    return result_rows, new_master, next_id, warnings


# ── Load Excel workbook data for one quarter ──────────────────────────────────

def load_excel_quarter(quarter):
    qdir = WORKBOOKS_DIR / quarter
    xlsxs = list(qdir.glob("*.xlsx")) if qdir.exists() else []
    if not xlsxs:
        return None, [], [], []
    wb = openpyxl.load_workbook(xlsxs[0], read_only=True, data_only=True)
    sheet = {s.lower(): s for s in wb.sheetnames}

    projects  = parse_excel_main(wb[sheet['main']], quarter) if 'main' in sheet else []
    load_chg  = parse_excel_load_changes(wb[sheet['load change']]) if 'load change' in sheet else []
    date_chg  = parse_excel_date_changes(wb[sheet['initial service change']]) if 'initial service change' in sheet else []
    ramp_chg  = parse_excel_ramp_changes(wb[sheet['ramp change']]) if 'ramp change' in sheet else []
    removed   = parse_excel_removed(wb[sheet['removed projects']]) if 'removed projects' in sheet else []
    return projects, load_chg, date_chg, ramp_chg, removed


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    COMBINED_DIR.mkdir(parents=True, exist_ok=True)

    master = {}       # {proj_id: record}
    next_id = 1
    all_rows = []

    for quarter in QUARTER_ORDER:
        print(f"\n[{quarter}]")

        if quarter == "2026Q1":
            projects  = parse_csv_main(quarter)
            load_chg  = parse_csv_load_changes()
            date_chg  = parse_csv_date_changes()
            ramp_chg  = []   # no ramp change CSV for 2026Q1
            removed   = parse_csv_removed()
        else:
            projects, load_chg, date_chg, ramp_chg, removed = load_excel_quarter(quarter)
            if projects is None:
                print(f"  Skipping — no workbook found")
                continue

        print(f"  {len(projects)} projects in Main sheet")
        print(f"  {len(load_chg)} load chg, {len(date_chg)} date chg, {len(ramp_chg)} ramp chg, {len(removed)} removals")

        if not master:
            # First quarter — assign IDs sequentially, no matching needed
            for proj in projects:
                pid = f"P{next_id:04d}"
                next_id += 1
                master[pid] = {
                    'segment':              proj['segment'],
                    'class_type':           proj['class_type'],
                    'territory':            proj['territory'],
                    'announced_load':       proj['announced_load'],
                    'initial_service_date': proj['initial_service_date'],
                    'ramp':                 proj['ramp'],
                    'quarter_first_seen':   quarter,
                }
                row = {
                    'proj_id': pid, 'report_quarter': quarter,
                    'quarter_first_seen': quarter, 'match_confidence': 'seed',
                    'segment': proj['segment'], 'class_type': proj['class_type'],
                    'territory': proj['territory'], 'project_stage': proj['project_stage'],
                    'announced_load_mw': proj['announced_load'],
                    'initial_service_date': proj['initial_service_date'],
                    'new_project': proj['new_project'],
                    'change_announced_load': proj['change_announced_load'],
                    'change_load_ramp': proj['change_load_ramp'],
                    'change_project_stage': proj['change_project_stage'],
                    'change_initial_service_date': proj['change_initial_service_date'],
                }
                for yr in LOAD_YEARS:
                    row[f'load_{yr}'] = proj['ramp'].get(yr) if yr in proj['ramp'] else None
                all_rows.append(row)
            print(f"  Seeded {len(projects)} projects (IDs P0001–P{next_id - 1:04d})")
        else:
            rows, master, next_id, warns = match_quarter(
                quarter, projects, master, load_chg, date_chg, ramp_chg, removed, next_id)
            all_rows.extend(rows)

            conf_counts = {}
            for r in rows:
                k = r['match_confidence'].split('_')[0]
                conf_counts[k] = conf_counts.get(k, 0) + 1
            print(f"  Match summary: {dict(sorted(conf_counts.items()))}")
            for w in warns:
                print(w)

    # Output
    df = pd.DataFrame(all_rows)

    # Add project_age: number of quarters this project has appeared in the pipeline
    df = df.sort_values(['proj_id', 'report_quarter'])
    df['project_age'] = df.groupby('proj_id').cumcount() + 1

    # Sort columns sensibly
    id_cols = ['proj_id', 'report_quarter', 'quarter_first_seen', 'project_age', 'match_confidence']
    attr_cols = ['segment', 'class_type', 'territory', 'project_stage',
                 'announced_load_mw', 'initial_service_date']
    flag_cols = ['new_project', 'change_announced_load', 'change_load_ramp',
                 'change_project_stage', 'change_initial_service_date']
    yr_cols = [f'load_{yr}' for yr in LOAD_YEARS]
    df = df[id_cols + attr_cols + flag_cols + yr_cols]

    out = COMBINED_DIR / "pipeline_projects.csv"
    df.to_csv(out, index=False)
    print(f"\nWrote {out}")
    print(f"  {len(df)} total rows, {df['proj_id'].nunique()} unique project IDs")

    # Summary by quarter
    print("\n--- Per-quarter summary ---")
    print(f"{'Quarter':<10} {'Projects':>9} {'New':>5} {'Exact':>7} {'Ramp':>6} {'Unmatched':>10}")
    print("-" * 52)
    for q in QUARTER_ORDER:
        sub = df[df.report_quarter == q]
        if sub.empty:
            continue
        new_ct  = (sub.match_confidence == 'new').sum() + (sub.match_confidence == 'seed').sum()
        exact   = sub.match_confidence.str.startswith('exact').sum()
        ramp    = sub.match_confidence.str.startswith('ramp').sum()
        unk     = len(sub) - new_ct - exact - ramp
        print(f"{q:<10} {len(sub):>9} {new_ct:>5} {exact:>7} {ramp:>6} {unk:>10}")


if __name__ == "__main__":
    main()
